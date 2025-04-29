import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import sys
import os
import re

# Добавляем родительскую директорию в путь для импорта
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from link_tables import apply_links

import plotly.express as px
import plotly.graph_objects as go
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

def load_data():
    """Загрузка и объединение данных"""
    # Загрузка основных данных
    df_orders = pd.read_excel('filtered_datasets/bbOrders_filtered.xlsx')
    df_models = pd.read_excel('filtered_datasets/uatModelsTS_filtered.xlsx')
    df_type_ts = pd.read_excel('filtered_datasets/uatTypeTS_filtered.xlsx')
    
    print("\nПроверка исходных данных:")
    print(f"Количество строк в заказах: {len(df_orders)}")
    print(f"Количество строк в моделях: {len(df_models)}")
    print(f"Количество строк в типах ТС: {len(df_type_ts)}")
    
    # Анализ заказов с "Not Information"
    not_info_orders = df_orders[df_orders['ТипТС'] == 'Not Information']
    print("\nАнализ заказов с 'Not Information' в колонке ТипТС:")
    print(f"Количество заказов с 'Not Information': {len(not_info_orders)}")
    print(f"Процент от общего количества заказов: {len(not_info_orders)/len(df_orders)*100:.2f}%")
    
    if len(not_info_orders) > 0:
        print("\nСтатистика по заказам с 'Not Information':")
        print(f"Среднее количество пассажиров: {not_info_orders['КоличествоПассажиров'].mean():.1f}")
        print(f"Максимальное количество пассажиров: {not_info_orders['КоличествоПассажиров'].max()}")
        print(f"Минимальное количество пассажиров: {not_info_orders['КоличествоПассажиров'].min()}")
        
        # Группировка по количеству пассажиров
        passenger_groups = not_info_orders.groupby(pd.cut(not_info_orders['КоличествоПассажиров'], 
                                                         bins=[0, 5, 10, 20, 50, float('inf')],
                                                         labels=['1-5', '6-10', '11-20', '21-50', '50+'])).size()
        print("\nРаспределение по количеству пассажиров:")
        for group, count in passenger_groups.items():
            print(f"{group}: {count} заказов ({count/len(not_info_orders)*100:.1f}%)")
    
    # Выводим первые несколько строк каждой таблицы для проверки
    print("\nПримеры данных из таблицы заказов (ТипТС):")
    print(df_orders[['ТипТС']].head())
    print("\nПримеры данных из таблицы типов ТС:")
    print(df_type_ts[['Ref', 'Description']].head())
    print("\nПримеры данных из таблицы моделей:")
    print(df_models[['Ref', 'Description', 'ВсегоМест', 'ТипТС']].head())
    
    # Выводим все уникальные значения Description из таблицы моделей
    print("\nВсе уникальные значения Description из таблицы моделей:")
    unique_descriptions = df_models['Description'].unique()
    for desc in sorted(unique_descriptions):
        if pd.notna(desc):  # Пропускаем NaN значения
            print(f"- {desc}")
    
    # Подготавливаем данные о вместимости
    # Группируем модели по типу ТС и берем максимальную вместимость
    capacity_by_type = df_models.groupby('ТипТС')['ВсегоМест'].max().reset_index()
    print("\nВместимость по типам ТС из таблицы моделей:")
    print(capacity_by_type.head())
    
    # Создаем словарь для хранения вместимости по типам ТС
    capacity_dict = dict(zip(capacity_by_type['ТипТС'], capacity_by_type['ВсегоМест']))
    
    # Добавляем вместимость для типов ТС, которых нет в таблице моделей
    # Извлекаем вместимость из названия типа ТС
    for ref, description in zip(df_type_ts['Ref'], df_type_ts['Description']):
        if ref not in capacity_dict:
            # Ищем число в начале названия (например, "51+1 ЮТОНГ Автобус")
            match = re.search(r'^(\d+)\+?\d*\s', description)
            if match:
                capacity = int(match.group(1))
                capacity_dict[ref] = capacity
                print(f"Добавлена вместимость для {description}: {capacity}")
    
    # Преобразуем словарь в DataFrame
    capacity_by_type = pd.DataFrame({
        'ТипТС': list(capacity_dict.keys()),
        'ВсегоМест': list(capacity_dict.values())
    })
    
    print("\nОбновленная вместимость по типам ТС:")
    print(capacity_by_type.head())
    
    # Объединяем данные
    # 1. Сначала объединяем заказы с типами ТС по Ref
    df_merged = pd.merge(
        df_orders,
        df_type_ts[['Ref', 'Description']].rename(columns={'Description': 'ТипТС_Description'}),
        left_on='ТипТС',
        right_on='Ref',
        how='left'
    )
    
    print("\nПроверка после первого объединения:")
    print(f"Количество строк после объединения с типами ТС: {len(df_merged)}")
    print(f"Количество уникальных значений ТипТС: {df_merged['ТипТС'].nunique()}")
    print(f"Количество уникальных значений ТипТС_Description: {df_merged['ТипТС_Description'].nunique()}")
    print("\nПримеры значений после первого объединения:")
    print(df_merged[['ТипТС', 'ТипТС_Description']].head())
    
    # 2. Затем объединяем с данными о вместимости
    df_merged = pd.merge(
        df_merged,
        capacity_by_type.rename(columns={'ТипТС': 'Ref', 'ВсегоМест': 'Вместимость'}),
        left_on='ТипТС',
        right_on='Ref',
        how='left'
    )
    
    print("\nПроверка после второго объединения:")
    print(f"Количество строк после объединения с данными о вместимости: {len(df_merged)}")
    print(f"Количество ненулевых значений Вместимость: {df_merged['Вместимость'].notna().sum()}")
    print("\nПримеры значений после второго объединения:")
    print(df_merged[['ТипТС', 'ТипТС_Description', 'Вместимость']].head())
    
    # Преобразуем числовые колонки
    df_merged['КоличествоПассажиров'] = pd.to_numeric(df_merged['КоличествоПассажиров'], errors='coerce')
    
    print("\nПроверка после преобразования типов:")
    print(f"Количество ненулевых значений Вместимость: {df_merged['Вместимость'].notna().sum()}")
    print(f"Количество ненулевых значений КоличествоПассажиров: {df_merged['КоличествоПассажиров'].notna().sum()}")
    
    return df_merged

def analyze_capacity(df):
    """Анализ вместимости ТС"""
    print("\nНачало анализа вместимости:")
    print(f"Количество строк для анализа: {len(df)}")
    
    # Группировка по типу ТС
    capacity_analysis = df.groupby('ТипТС_Description').agg({
        'Вместимость': 'first',
        'КоличествоПассажиров': ['count', 'mean', 'max', 'min'],
        'Number': 'count'
    }).reset_index()
    
    print("\nРезультаты группировки:")
    print(f"Количество групп: {len(capacity_analysis)}")
    print("\nПервые 5 строк результата:")
    print(capacity_analysis.head())
    
    # Переименование колонок
    capacity_analysis.columns = [
        'ТипТС', 'Вместимость', 'КоличествоЗаказов', 
        'СреднееКоличествоПассажиров', 'МаксКоличествоПассажиров',
        'МинКоличествоПассажиров', 'ВсегоЗаказов'
    ]
    
    # Добавляем процент использования от максимальной вместимости
    capacity_analysis['ПроцентИспользования'] = (
        capacity_analysis['СреднееКоличествоПассажиров'] / 
        capacity_analysis['Вместимость'] * 100
    ).round(2)
    
    # Сортируем по количеству заказов
    capacity_analysis = capacity_analysis.sort_values('КоличествоЗаказов', ascending=False)
    
    print("\nФинальные результаты:")
    print(f"Количество строк в финальном результате: {len(capacity_analysis)}")
    print("\nПервые 5 строк финального результата:")
    print(capacity_analysis.head())
    
    return capacity_analysis

def create_visualizations(df, capacity_analysis):
    """Создание визуализаций"""
    # Создаем директорию для графиков, если её нет
    if not os.path.exists('visualizations'):
        os.makedirs('visualizations')
    
    # 1. Распределение вместимости
    plt.figure(figsize=(12, 6))
    sns.histplot(data=capacity_analysis, x='Вместимость', bins=20)
    plt.title('Распределение вместимости ТС')
    plt.xlabel('Вместимость (мест)')
    plt.ylabel('Количество типов ТС')
    plt.savefig('visualizations/capacity_distribution.png')
    plt.close()
    
    # 2. Топ-10 самых используемых ТС
    plt.figure(figsize=(12, 6))
    top_10 = capacity_analysis.head(10)
    sns.barplot(data=top_10, x='ТипТС', y='КоличествоЗаказов')
    plt.title('Топ-10 самых используемых типов ТС')
    plt.xlabel('Тип ТС')
    plt.ylabel('Количество заказов')
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.savefig('visualizations/top_10_vehicles.png')
    plt.close()
    
    # 3. Соотношение вместимости и использования
    fig = px.scatter(
        capacity_analysis,
        x='Вместимость',
        y='СреднееКоличествоПассажиров',
        size='КоличествоЗаказов',
        hover_data=['ТипТС', 'ПроцентИспользования'],
        title='Соотношение вместимости и фактического использования'
    )
    fig.write_html('visualizations/capacity_usage.html')
    
    # 4. Распределение количества пассажиров
    plt.figure(figsize=(12, 6))
    sns.histplot(data=df, x='КоличествоПассажиров', bins=30)
    plt.title('Распределение количества пассажиров')
    plt.xlabel('Количество пассажиров')
    plt.ylabel('Количество заказов')
    plt.savefig('visualizations/passengers_distribution.png')
    plt.close()

def save_to_excel(capacity_analysis):
    """Сохранение результатов в Excel"""
    # Создаем директорию для результатов, если её нет
    if not os.path.exists('analysis_results'):
        os.makedirs('analysis_results')
    
    # Создаем Excel файл
    writer = pd.ExcelWriter('analysis_results/vehicle_capacity_analysis.xlsx', engine='openpyxl')
    capacity_analysis.to_excel(writer, sheet_name='Анализ вместимости', index=False)
    
    # Получаем рабочую книгу и лист
    workbook = writer.book
    worksheet = writer.sheets['Анализ вместимости']
    
    # Настраиваем стили
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    # Применяем стили к заголовкам
    for col in range(1, len(capacity_analysis.columns) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
    
    # Настраиваем ширину колонок
    for col in range(1, len(capacity_analysis.columns) + 1):
        column_letter = get_column_letter(col)
        worksheet.column_dimensions[column_letter].width = 20
    
    # Сохраняем файл
    writer.close()

def print_detailed_analysis(capacity_analysis):
    """Вывод подробного анализа в консоль"""
    print("\n" + "="*80)
    print("ПОДРОБНЫЙ АНАЛИЗ ИСПОЛЬЗОВАНИЯ ТРАНСПОРТНЫХ СРЕДСТВ")
    print("="*80)
    
    # 1. Топ-5 самых эффективных ТС (по проценту использования)
    print("\n1. Топ-5 ТС по эффективности использования вместимости:")
    print("-"*80)
    top_efficient = capacity_analysis[capacity_analysis['Вместимость'].notna()].sort_values('ПроцентИспользования', ascending=False).head()
    for _, row in top_efficient.iterrows():
        print(f"- {row['ТипТС']}:")
        print(f"  Вместимость: {row['Вместимость']} мест")
        print(f"  Среднее количество пассажиров: {row['СреднееКоличествоПассажиров']:.1f}")
        print(f"  Процент использования: {row['ПроцентИспользования']}%")
        print(f"  Количество заказов: {row['КоличествоЗаказов']}")
    
    # 2. Топ-5 по количеству заказов
    print("\n2. Топ-5 самых востребованных ТС:")
    print("-"*80)
    top_ordered = capacity_analysis.head()
    for _, row in top_ordered.iterrows():
        print(f"- {row['ТипТС']}:")
        print(f"  Количество заказов: {row['КоличествоЗаказов']}")
        print(f"  Вместимость: {row['Вместимость']} мест")
        print(f"  Среднее количество пассажиров: {row['СреднееКоличествоПассажиров']:.1f}")
        print(f"  Процент использования: {row['ПроцентИспользования']}%")
    
    # 3. Статистика по категориям вместимости
    print("\n3. Статистика по категориям вместимости:")
    print("-"*80)
    
    def get_category(capacity):
        if pd.isna(capacity):
            return "Нет данных"
        elif capacity <= 8:
            return "Малая (до 8 мест)"
        elif capacity <= 20:
            return "Средняя (9-20 мест)"
        elif capacity <= 35:
            return "Большая (21-35 мест)"
        else:
            return "Особо большая (36+ мест)"
    
    capacity_analysis['Категория'] = capacity_analysis['Вместимость'].apply(get_category)
    category_stats = capacity_analysis.groupby('Категория').agg({
        'КоличествоЗаказов': 'sum',
        'ПроцентИспользования': 'mean'
    }).reset_index()
    
    for _, row in category_stats.iterrows():
        print(f"- {row['Категория']}:")
        print(f"  Всего заказов: {row['КоличествоЗаказов']}")
        if not pd.isna(row['ПроцентИспользования']):
            print(f"  Средний процент использования: {row['ПроцентИспользования']:.1f}%")
    
    print("\n" + "="*80)

def main():
    # Загрузка данных
    print("Загрузка данных...")
    df = load_data()
    
    # Анализ вместимости
    print("Анализ вместимости...")
    capacity_analysis = analyze_capacity(df)
    
    # Вывод подробного анализа
    print_detailed_analysis(capacity_analysis)
    
    # Создание визуализаций
    print("Создание визуализаций...")
    create_visualizations(df, capacity_analysis)
    
    # Сохранение результатов
    print("Сохранение результатов...")
    save_to_excel(capacity_analysis)
    
    print("Анализ завершен!")

if __name__ == "__main__":
    main()
